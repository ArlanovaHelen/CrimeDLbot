from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sqlite_db
def find_sql(text):
    first = text.index("SELECT")
    second = text.index("```", first)

    final = text[first:second]
    return final





def create_statistic_file(data):
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(list(row))
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="AFE1AF", end_color="AFE1AF", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        ws.cell(row=1, column=col_num).font = header_font
        ws.cell(row=1, column=col_num).alignment = alignment_center
        ws.cell(row=1, column=col_num).fill = header_fill
        ws.cell(row=1, column=col_num).border = border


    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save('report.xlsx')
    wb.close()



